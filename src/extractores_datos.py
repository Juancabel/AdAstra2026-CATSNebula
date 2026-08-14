"""
extractores_datos.py — Extracción de texto para los formatos de B.

Formatos: json, xlsx, imagen (OCR), pbf (Mapbox Vector Tiles).

Cada extractor devuelve una tupla (texto, titulo, extra):
    texto  : str  — el contenido que se va a indexar
    titulo : str | None
    extra  : dict — metadata barata, lo que sale gratis del parseo

Ninguno lanza excepción por contenido vacío: eso lo decide ingest_data.py
al llamar a compute_doc_id().
"""

import collections
import csv
import json
from datetime import date, datetime
from pathlib import Path
import re

from mapeos_json import (
    ARCHIVOS_EXCLUIDOS,
    ATRIBUTOS_PBF_DESCARTADOS,
    MAPEO_ALERTAS,
    ATRIBUTOS_PBF_PRIORITARIOS,
    ATRIBUTOS_PBF_SEMANTICOS,
    CAMPOS_RUIDO_CATALOGO,
    FAMILIA_ARTICULO_WEB,
    MAPEOS_POR_INSTITUCION,
    PATRONES_CATALOGO,
)


# ---------------------------------------------------------------------------
# Evaluador de rutas de campo
# ---------------------------------------------------------------------------

def extraer_ruta(obj, ruta: str) -> list[str]:
    """
    Evalúa una ruta de campo sobre un objeto JSON y devuelve la lista de
    strings encontrados.

    Soporta:
        "title"                    -> clave directa
        "body_paragraphs[]"        -> lista de strings
        "alerta_meta.tema_clave"   -> clave anidada
        "sections[].paragraphs[]"  -> anidamiento con listas

    Args:
        obj: el objeto JSON (o subobjeto) sobre el que evaluar.
        ruta: expresión de ruta. Cadena vacía significa "este nodo es la hoja".

    Returns:
        Lista de strings no vacíos. Lista vacía si la ruta no existe.
    """
    if obj is None:
        return []

    # Caso hoja: no queda ruta por recorrer.
    if not ruta:
        if isinstance(obj, str):
            return [obj] if obj.strip() else []
        if isinstance(obj, bool):
            return []  # los booleanos sueltos no aportan nada al texto
        if isinstance(obj, (int, float)):
            return [formatear_valor(obj)]
        if isinstance(obj, list):
            resultado = []
            for item in obj:
                resultado.extend(extraer_ruta(item, ""))
            return resultado
        return []

    partes = ruta.split(".", 1)
    segmento = partes[0]
    resto = partes[1] if len(partes) > 1 else ""

    es_lista = segmento.endswith("[]")
    clave = segmento[:-2] if es_lista else segmento

    if not isinstance(obj, dict):
        return []

    valor = obj.get(clave)
    if valor is None:
        return []

    if es_lista:
        if not isinstance(valor, list):
            valor = [valor]
        resultado = []
        for item in valor:
            resultado.extend(extraer_ruta(item, resto))
        return resultado

    return extraer_ruta(valor, resto)

_VINETA = re.compile(r"^\s*[-•*\u2022\u2013\u2014]\s+")


def aplanar_valor(texto: str) -> str:
    """
    Colapsa un valor multilínea en UNA sola línea.

    En datos de registro la unidad atómica de fragmentación es la línea: un
    valor con saltos internos parte el registro en trozos huérfanos que el
    chunker trataría como registros independientes, incumpliendo el §3.3.

    Distingue dos casos porque el separador correcto no es el mismo:
      - LISTA    (alguna línea empieza por viñeta)  -> se une con "; "
      - ENVUELTO (texto partido por ancho de celda) -> se une con " "

    Se usa SOLO en contextos de registro (xlsx, pbf, catálogos json).
    NUNCA sobre body_paragraphs: ahí los \\n\\n son estructura de párrafo.
    """
    if not texto:
        return ""
    lineas = [l for l in texto.splitlines() if l.strip()]
    if not lineas:
        return ""
    es_lista = any(_VINETA.match(l) for l in lineas)
    limpias = [x for x in (_VINETA.sub("", l).strip() for l in lineas) if x]
    plano = ("; " if es_lista else " ").join(limpias)
    plano = plano.replace("|", "/")
    return re.sub(r"[ \t\u00a0]+", " ", plano).strip()


def formatear_valor(v) -> str:
    """
    Convierte un valor a texto de forma limpia.

    Resuelve la trampa confirmada en el reconocimiento: openpyxl y json
    devuelven enteros como float (32634855.0 en vez de 32634855), y en
    texto indexado eso es peor que el entero.
    """
    if v is None:
        return ""
    if isinstance(v, bool):
        return "sí" if v else "no"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    return str(v)


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

def es_catalogo(ruta: Path) -> bool:
    """Detecta catálogos de scraping por el nombre del archivo."""
    nombre = ruta.name.lower()
    return any(p in nombre for p in PATRONES_CATALOGO)


def extraer_catalogo(datos) -> str:
    """
    Extrae solo los campos semánticos de un catálogo de scraping.

    Descarta URLs, tamaños, timestamps, hashes y códigos internos: son
    ruido que compite por posiciones en el top-10 sin poder responder
    nunca a una pregunta en lenguaje natural.

    Usa LISTA NEGRA, no blanca: en caso de duda se indexa. Una lista blanca
    descartaba contenido real (p.ej. el campo "valor" de SWF_report-data.json).
    """
    registros = datos if isinstance(datos, list) else [datos]
    lineas = []

    def _collect_strings(obj):
        """Recolecta recursivamente cadenas y valores simples de un objeto.

        Devuelve una lista de strings representando los valores encontrados.
        """
        res = []
        if obj is None:
            return res
        if isinstance(obj, str):
            if obj.strip():
                res.append(obj)
            return res
        if isinstance(obj, (int, float, bool)):
            s = formatear_valor(obj)
            if s:
                res.append(s)
            return res
        if isinstance(obj, dict):
            for v in obj.values():
                res.extend(_collect_strings(v))
            return res
        if isinstance(obj, list):
            for item in obj:
                res.extend(_collect_strings(item))
            return res
        return res

    for registro in registros:
        if not isinstance(registro, dict):
            continue
        partes = []
        for clave, valor in registro.items():
            clave_norm = clave.lower().strip()
            if clave_norm in CAMPOS_RUIDO_CATALOGO:
                continue

            # Si el valor es un diccionario grande (p.ej. 'articulos': {url: path}),
            # extraer sus cadenas internas en lugar de intentar evaluar la ruta.
            if isinstance(valor, dict) or isinstance(valor, list):
                textos = _collect_strings(valor)
            else:
                textos = extraer_ruta({clave: valor}, clave)

            if textos:
                # aplanar_valor: un catálogo es dato de registro (una línea por
                # registro), así que ningún valor puede traer saltos internos.
                partes.append(f"{clave}: {aplanar_valor(' '.join(textos))}")
        if partes:
            lineas.append(" | ".join(partes))

    return "\n".join(lineas)


def componer_titulo_alerta(datos: dict) -> str | None:
    """
    Compone un título sintético para Alertas Tempranas.

    Los 363 archivos traen title="Mapa", que es inútil. Un título como
    "Alerta 001-17 — Cartagena de Indias (Bolívar)" es mucho más útil
    para el experimento de prefijo de título del Día 5.
    """
    meta = datos.get("alerta_meta") or {}
    codigo = meta.get("codigo")
    municipios = meta.get("municipios")
    if codigo and municipios:
        return f"Alerta {codigo} — {municipios}"
    if codigo:
        return f"Alerta {codigo}"
    return None


# --- Pies de plantilla en los JSON de artículo web --------------------------
#
# Los scrapers dejaron dentro de `body_paragraphs` el pie de la página: el
# "About us" del observatorio, el copyright, la dirección postal, el "Related
# Experts:". Queda pegado al final del último párrafo real y produce chunks que
# empiezan a mitad de una oración ajena.
#
# Igual que con los PDF: se detecta por REPETICIÓN entre artículos del mismo
# sitio, no por palabras clave — así no depende del idioma ni del formato
# exacto del byline.
MIN_ARTICULOS_PIE = 3

# Palabras de la firma por cada extremo. Comparar solo las primeras y últimas
# permite que cambie el nombre del autor en mitad del pie.
PALABRAS_FIRMA = 6


def _firma_parrafo(parrafo: str) -> tuple[str, str]:
    """Firma laxa de un párrafo: primeras y últimas PALABRAS_FIRMA palabras."""
    palabras = " ".join(parrafo.split()).lower().split()
    if not palabras:
        return ("", "")
    return (" ".join(palabras[:PALABRAS_FIRMA]), " ".join(palabras[-PALABRAS_FIRMA:]))


def _parrafos_de_articulo(datos, mapeo) -> list[str]:
    """Los párrafos que el mapeo extraería de este JSON, en orden."""
    objetos = ([d for d in datos if isinstance(d, dict)]
               if isinstance(datos, list) else [datos])
    parrafos = []
    for obj in objetos:
        if not isinstance(obj, dict):
            continue
        for campo in mapeo["texto"]:
            parrafos.extend(t for t in extraer_ruta(obj, campo) if t.strip())
    return parrafos


def detectar_pies_plantilla(raiz: Path) -> dict[str, set]:
    """
    Encuentra los pies de plantilla de cada observatorio de artículo web.

    Se calcula UNA vez sobre todo el corpus antes de la ingesta, igual que
    dedup_pbf.construir_asignacion: la repetición solo se ve mirando varios
    artículos a la vez, y un extractor que procesa un archivo no puede verla.

    Un párrafo es pie de plantilla si cumple LAS DOS condiciones:

      1. cierra al menos MIN_ARTICULOS_PIE artículos del mismo observatorio, y
      2. NO aparece nunca a mitad del cuerpo de ningún artículo del sitio.

    La segunda no es un adorno. Sin ella se borraba contenido real: en
    SWF_Counterspace el párrafo "Since 2010, Russia has been testing..." cierra
    9 artículos, pero es un resumen que el observatorio reutiliza, y en otros
    artículos aparece en mitad del texto. Un pie de plantilla, por definición,
    nunca cae a mitad de cuerpo; un párrafo reutilizado sí.

    Returns:
        {institucion: {firma, ...}} — determinista, sin estado entre corridas.
    """
    finales = collections.defaultdict(collections.Counter)
    interiores = collections.defaultdict(set)

    for ruta in sorted(raiz.rglob("*.json")):
        ruta_relativa = ruta.relative_to(raiz).as_posix()
        partes = ruta_relativa.split("/")
        if len(partes) < 2:
            continue
        institucion = "/".join(partes[:2])
        mapeo = MAPEOS_POR_INSTITUCION.get(institucion)
        if mapeo is not FAMILIA_ARTICULO_WEB:
            continue
        if ruta.name in ARCHIVOS_EXCLUIDOS or es_catalogo(ruta):
            continue

        try:
            datos = json.loads(ruta.read_text(encoding="utf-8", errors="replace"))
        except (json.JSONDecodeError, OSError):
            continue

        parrafos = _parrafos_de_articulo(datos, mapeo)
        if len(parrafos) < 2:
            continue

        finales[institucion][_firma_parrafo(parrafos[-1])] += 1
        for p in parrafos[:-1]:
            interiores[institucion].add(_firma_parrafo(p))

    plantillas = {}
    for institucion, cuenta in finales.items():
        detectadas = {
            firma for firma, n in cuenta.items()
            if n >= MIN_ARTICULOS_PIE
            and firma not in interiores[institucion]
            and any(firma)
        }
        if detectadas:
            plantillas[institucion] = detectadas
    return plantillas


def es_indice_teselas(ruta: Path) -> bool:
    """Detecta el manifiesto de descarga de teselas (AMAZONUW_tiles-index.json)."""
    return "tiles-index" in ruta.name.lower()


def extraer_indice_teselas(datos) -> tuple[str, str | None, dict]:
    """
    Resume el manifiesto de descarga de teselas como texto de cobertura.

    POR QUÉ NO SE SERIALIZA COMO CATÁLOGO
    Sus once campos (tile, zoom, x, y, url, local_path, status, size_bytes,
    content_type, from_cache, error) son SIN EXCEPCIÓN contabilidad del
    scraping: están todos en CAMPOS_RUIDO_CATALOGO, y por eso el catálogo
    genérico lo dejaba vacío. No hay un solo topónimo ni descripción.

    Pero el archivo tiene DOC_ID oficial, así que no puede quedar fuera del
    entregable. Emitir sus 262 filas de coordenadas y códigos HTTP metería en
    el índice justo el ruido que la lista negra existe para evitar. La salida
    intermedia es un resumen de cobertura: dice qué mapa describe y qué
    porción se pudo descargar, que es la única información recuperable aquí.

    Es el mismo recurso que dedup_pbf usa para las teselas cuyos features ya
    se indexaron en otro nivel de zoom (CLAVE_RESUMEN).
    """
    registros = [r for r in (datos if isinstance(datos, list) else [datos])
                 if isinstance(r, dict)]
    total = len(registros)
    con_datos = sorted(r.get("tile") for r in registros
                       if str(r.get("status")) == "cached")
    sin_datos = total - len(con_datos)
    zooms = sorted({r["zoom"] for r in registros if isinstance(r.get("zoom"), int)})

    lineas = [
        "Índice de teselas del mapa vectorial de Amazon Underworld.",
        f"teselas solicitadas: {total} | con datos: {len(con_datos)} | "
        f"sin datos: {sin_datos}",
    ]
    if zooms:
        lineas.append(
            f"niveles de zoom cubiertos: {zooms[0]} a {zooms[-1]} "
            f"({len(zooms)} niveles)"
        )
    if con_datos:
        lineas.append("teselas con datos: " + ", ".join(con_datos))

    extra = {
        "teselas_totales": total,
        "teselas_con_datos": len(con_datos),
        "teselas_sin_datos": sin_datos,
        "zooms": zooms,
    }
    return "\n".join(lineas), "Índice de teselas — Amazon Underworld", extra


def extraer_json(ruta: Path, institucion: str,
                 pies_plantilla: set | None = None) -> tuple[str, str | None, dict]:
    """
    Extrae texto de un archivo JSON según el mapeo de su institución.

    Args:
        ruta: ruta al archivo.
        institucion: clave de MAPEOS_POR_INSTITUCION.
        pies_plantilla: firmas de pie de este observatorio, de
            detectar_pies_plantilla(). Los párrafos finales que casen se
            descartan. Sin este argumento no se limpia nada.

    Returns:
        (texto, titulo, extra)

    Raises:
        ValueError: si la institución no tiene mapeo definido. Es
            intencional: adivinar campos produce documentos silenciosamente
            vacíos o llenos de URLs. Mejor fallar y añadir el mapeo.
    """
    contenido = ruta.read_text(encoding="utf-8", errors="replace")
    datos = json.loads(contenido)

    # El manifiesto de teselas va antes que el catálogo genérico: encaja en
    # PATRONES_CATALOGO ("tiles-index") pero todos sus campos son ruido, así
    # que el catálogo lo dejaría vacío.
    if es_indice_teselas(ruta):
        return extraer_indice_teselas(datos)

    # Los catálogos se detectan por nombre de archivo, antes que por institución.
    if es_catalogo(ruta):
        return extraer_catalogo(datos), None, {}

    mapeo = MAPEOS_POR_INSTITUCION.get(institucion)
    if mapeo is None:
        raise ValueError(
            f"Institución sin mapeo definido: {institucion!r}. "
            f"Añadirla a MAPEOS_POR_INSTITUCION en mapeos_json.py."
        )

    if isinstance(datos, list):
        # Un JSON con raíz de lista que no es catálogo: se concatenan los
        # elementos. El spec §1.3 dice archivo = documento.
        objetos = [d for d in datos if isinstance(d, dict)]
    else:
        objetos = [datos]

    bloques = []
    for obj in objetos:
        for campo in mapeo["texto"]:
            textos = extraer_ruta(obj, campo)
            bloques.extend(textos)

    bloques = [b.strip() for b in bloques if b.strip()]

    # Pie de plantilla: se recortan los párrafos finales que coincidan, no solo
    # el último — hay sitios que encadenan dos (copyright + "Related Experts").
    # Solo desde el final: la misma firma a mitad de cuerpo es contenido, y
    # detectar_pies_plantilla() ya descartó las firmas que aparecen ahí.
    pies_recortados = 0
    if pies_plantilla:
        while len(bloques) > 1 and _firma_parrafo(bloques[-1]) in pies_plantilla:
            bloques.pop()
            pies_recortados += 1

    texto = "\n\n".join(bloques)

    # Red de seguridad: si el mapeo de la institución no encontró nada, el
    # archivo puede tener una estructura distinta a la del resto de su fuente
    # (p.ej. SWF_report-data.json usa seccion/campo/valor y no title/body).
    # Antes de darlo por perdido, se intenta extracción genérica por lista
    # negra. Vale más indexar algo imperfecto que perder el documento: si el
    # ground truth lo referencia y no está indexado, es F1@3 perdido.
    if not texto.strip():
        texto = extraer_catalogo(datos)

    # Título
    titulo = None
    if mapeo is MAPEO_ALERTAS and objetos:
        titulo = componer_titulo_alerta(objetos[0])
    elif mapeo.get("titulo") and objetos:
        candidatos = extraer_ruta(objetos[0], mapeo["titulo"])
        titulo = candidatos[0] if candidatos else None

    # Extra: lo que sale gratis del parseo
    extra = {}
    if objetos:
        for campo in mapeo.get("extra", []):
            valores = extraer_ruta(objetos[0], campo)
            if valores:
                extra[campo.split(".")[-1]] = valores[0]
    if pies_recortados:
        extra["pies_plantilla_recortados"] = pies_recortados

    return texto, titulo, extra


# ---------------------------------------------------------------------------
# Texto plano / HTML / Markdown / PDF
# ---------------------------------------------------------------------------

def extraer_txt(ruta: Path) -> tuple[str, str | None, dict]:
    """Lee un archivo de texto plano y lo devuelve como contenido.

    Devuelve (texto, titulo=None, extra={}). Usa 'replace' en errores de
    decodificación para evitar fallos por codificaciones heterogéneas.
    """
    contenido = ruta.read_text(encoding="utf-8", errors="replace")
    return contenido, None, {}


def extraer_md(ruta: Path) -> tuple[str, str | None, dict]:
    """Extrae texto desde Markdown: limpia sintaxis básica.

    Mantiene el texto visible y convierte enlaces `[a](url)` a `a`.
    """
    s = ruta.read_text(encoding="utf-8", errors="replace")
    # Quitar code fences
    s = re.sub(r"```.*?```", "", s, flags=re.S)
    # Reemplazar enlaces [texto](url) -> texto
    s = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", s)
    # Quitar imágenes ![alt](url) -> alt
    s = re.sub(r"!\[([^\]]*)\]\([^\)]+\)", r"\1", s)
    # Quitar inline code `code`
    s = re.sub(r"`([^`]+)`", r"\1", s)
    # Quitar encabezados markdown
    s = re.sub(r"^#{1,6}\s*", "", s, flags=re.M)
    return s, None, {}


def extraer_html(ruta: Path) -> tuple[str, str | None, dict]:
    """Extrae texto desde HTML sin depender de BeautifulSoup.

    Elimina scripts/estilos y etiquetas, y desempaqueta entidades HTML.
    No es perfecto pero suficiente como extractor de fallback.
    """
    from html import unescape

    s = ruta.read_text(encoding="utf-8", errors="replace")
    # Eliminar scripts y estilos
    s = re.sub(r"(?is)<script.*?>.*?</script>", "", s)
    s = re.sub(r"(?is)<style.*?>.*?</style>", "", s)
    # Reemplazar etiquetas por espacio
    s = re.sub(r"(?s)<[^>]+>", " ", s)
    s = unescape(s)
    s = re.sub(r"[ \t\xa0]+", " ", s)
    s = s.strip()
    return s, None, {}


# --- Boilerplate repetido de los PDF ----------------------------------------
#
# El 73% de los chunks de PDF no terminaban en signo de cierre. No era culpa
# del chunker: le llegaban unidades que mezclaban el final de una oración con
# el encabezado de la página siguiente, que PyMuPDF emite como texto más.
#
# El principio es el mismo que ya se usó en dedup_pbf: DETECTAR REPETICIÓN, no
# adivinar contenido. No hay lista de palabras prohibidas; se elimina lo que se
# repite tanto entre páginas que no puede ser el cuerpo del documento.
#
# Medido sobre la muestra de la Fase 0 (6 PDF, de 51 a 1.330 páginas), la
# primera línea normalizada de cada página se repite en el 82%, 99,9%, 98%,
# 99,8%, 49%+46% y 41%+39% de las páginas respectivamente. Los dos últimos
# alternan encabezado par/impar, que es justo lo que obliga a que el umbral
# sea bajo: con 50% se escaparían.
UMBRAL_REPETICION_PAGINA = 0.30

# Ventana de líneas candidatas por página. El encabezado NO es una línea: en
# AI_Index son seis consecutivas ("Artificial Intelligence" / "Index Report
# 2024" / "CHAPTER 1:" / "Research and" / "Development" + numeración). Mirar
# solo la primera línea dejaba intacto justo el boilerplate del caso original.
# 8/4 da margen sobre ese bloque de 6 sin arriesgar: ampliar la ventana amplía
# qué se EVALÚA, no qué se borra — el umbral del 30% sigue protegiendo al
# contenido único.
VENTANA_CABECERA = 8
VENTANA_PIE = 4

# Salvaguarda: si la limpieza se llevaría más de esta fracción de las palabras
# del documento, NO se aplica y el documento se marca para revisión manual.
# Mismo espíritu que "conservar metadata sin inventar contenido": mejor una
# nota de cobertura que un documento vaciado por un patrón mal detectado.
MAX_FRACCION_ELIMINABLE = 0.40

_RE_DIGITOS = re.compile(r"\d+")

# Numeración de página: solo dígitos, o "Page 4", "4 / 120", "- 12 -".
_RE_NUMERACION = re.compile(
    r"^(?:p(?:age|ág?\.?|\.)\s*)?[-–—\s]*\d+(?:\s*(?:/|\||de|of)\s*\d+)?[-–—\s]*$",
    re.IGNORECASE,
)


def _es_numeracion(linea: str) -> bool:
    """La línea es solo un número de página, con o sin adornos."""
    return bool(_RE_NUMERACION.match(linea.strip()))


def _normalizar_linea(linea: str) -> str:
    """
    Forma canónica para comparar líneas entre páginas.

    Los dígitos pasan a '#' para que "Page 4" y "Page 5" cuenten como la misma
    línea; sin eso ningún encabezado con numeración se detectaría nunca.

    El precio conocido: en AI_Index los rótulos "Figure 1.3.1" y "Figure 1.3.4"
    colapsan en "Figure #.#.#" y se van con el boilerplate (20 líneas de tres
    palabras, ya separadas de su figura por el orden de extracción). Se acepta:
    la alternativa —exigir que la forma cruda también se repita— desactivaría
    justo el caso que hay que cazar.
    """
    return _RE_DIGITOS.sub("#", " ".join(linea.split())).strip()


def _indices_ventana(n: int) -> set[int]:
    """Posiciones de las primeras VENTANA_CABECERA y últimas VENTANA_PIE."""
    return set(range(min(VENTANA_CABECERA, n))) | set(
        range(max(0, n - VENTANA_PIE), n)
    )


def limpiar_boilerplate_paginas(paginas: list[str]) -> tuple[list[str], dict]:
    """
    Quita numeración y encabezados/pies repetidos de un PDF ya paginado.

    Args:
        paginas: texto de cada página, en orden.

    Returns:
        (paginas_limpias, diagnostico). El diagnóstico lleva qué se eliminó y
        cuánto, para que ingest_data.py lo registre en el log: una limpieza
        silenciosa es tan mala como el boilerplate.
    """
    # Cada página como lista de líneas no vacías.
    por_pagina = [
        [l.strip() for l in pagina.splitlines() if l.strip()] for pagina in paginas
    ]
    con_texto = [p for p in por_pagina if p]
    if len(con_texto) < 2:
        # Sin al menos dos páginas no hay repetición que medir.
        return paginas, {"lineas_eliminadas": 0}

    # PASO 1 — Numeración de página, antes que nada. Se descarta sin exigir
    # repetición, pero SOLO dentro de la ventana: un "2010" suelto en mitad de
    # la página es la etiqueta del eje de una gráfica, no un número de página.
    numeracion_fuera = 0
    sin_numeracion = []
    for lineas in por_pagina:
        ventana = _indices_ventana(len(lineas))
        conservadas = []
        for i, linea in enumerate(lineas):
            if i in ventana and _es_numeracion(linea):
                numeracion_fuera += 1
                continue
            conservadas.append(linea)
        sin_numeracion.append(conservadas)

    # PASO 2 — Recalcular la ventana ya sin numeración: así el encabezado real
    # ocupa las 8 posiciones y no las desperdicia en el número de página.
    conteo = collections.Counter()
    for lineas in sin_numeracion:
        if not lineas:
            continue
        ventana = _indices_ventana(len(lineas))
        # set(): una línea repetida dos veces en la MISMA página cuenta una vez.
        for forma in {_normalizar_linea(lineas[i]) for i in ventana}:
            if forma:
                conteo[forma] += 1

    n_paginas = sum(1 for p in sin_numeracion if p)
    minimo = max(2, int(n_paginas * UMBRAL_REPETICION_PAGINA))
    repetidas = {forma for forma, n in conteo.items() if n >= minimo}

    # PASO 3 — Eliminar las repetidas, solo donde aparecen en la ventana. Fuera
    # de ella la misma cadena puede ser contenido legítimo: "Development" es
    # encabezado arriba y palabra normal en mitad de un párrafo.
    limpias = []
    eliminadas = 0
    for lineas in sin_numeracion:
        if not lineas:
            limpias.append("")
            continue
        ventana = _indices_ventana(len(lineas))
        conservadas = []
        for i, linea in enumerate(lineas):
            if i in ventana and _normalizar_linea(linea) in repetidas:
                eliminadas += 1
                continue
            conservadas.append(linea)
        limpias.append("\n".join(conservadas))

    diagnostico = {
        "lineas_eliminadas": eliminadas + numeracion_fuera,
        "numeracion_eliminada": numeracion_fuera,
        "encabezados_pies_eliminados": eliminadas,
        "patrones_detectados": sorted(repetidas)[:10],
        "n_patrones": len(repetidas),
    }
    return limpias, diagnostico


# --- Líneas que no son prosa, en mitad de la página -------------------------
#
# APAGADO POR DEFECTO (`_aplicar_limpieza(aislar=False)`). MEDIDO Y NO PAGA:
# sobre la muestra de 19 documentos de scripts/ab_extraccion.py, aislar 1.185
# líneas mueve UN chunk de 433 (corte_real 42,9% -> 42,7%) y empeora tres
# documentos. No justifica una reingesta de 70 min ni invalidar la caché de
# chunking. El motivo es estructural y se explica al final de este bloque.
# Se deja implementado y apagado para que la medición sea reproducible y para
# que activarlo sea un flag, no volver a escribirlo.
#
# El limpiador de arriba solo mira la ventana de cabecera/pie y solo borra lo
# que se REPITE. Quedan fuera dos cosas medidas en la muestra:
#
#   1. Encabezados de sección con alcance de sección, no de documento:
#      "1.1 Publications" sale en 8 de las 51 páginas de AIINDEX-001 y "1.2
#      Patents" en otras 8. Están en la ventana, pero ninguno llega al 30% de
#      UMBRAL_REPETICION_PAGINA, así que sobreviven. Bajar el umbral no es la
#      salida: se llevaría contenido único por delante.
#   2. Etiquetas de datos de gráficas: "0.05, Clinical trial", "1.46%, Other".
#      No las caza clasificar_residual.py porque no llevan "Chart:"/"Source:"
#      ni cola numérica pura.
#
# LO QUE SE HACE NO ES BORRAR, ES SEPARAR. Un encabezado de sección es
# contenido legítimo y recuperable; el problema no es que esté, es que llega
# pegado al final de la oración anterior — "…conference papers. 1.1
# Publications Total Number of AI Publications" — y el chunker, que nunca
# parte una unidad, se lo come entero.
#
# El texto de las páginas se une con "\n\n" y chunk.py parte párrafos justo
# ahí (`re.split(r"\n\s*\n")`), así que rodear la línea de líneas en blanco
# la convierte en su propia unidad. Cero contenido perdido, y un falso
# positivo cuesta un salto de párrafo de más, no un borrado.
#
# El coste de equivocarse es asimétrico y por eso las guardas son estrictas:
# aislar una línea que SÍ era prosa parte una oración en dos, que es
# exactamente el defecto que se quiere arreglar. De ahí las dos condiciones de
# contexto: solo se aísla si la línea anterior YA cerró oración (no se corta
# nada a medias) y si la siguiente no es su continuación en minúscula.
#
# POR QUÉ NO MOVIÓ LA MÉTRICA (lo que enseñó la medición):
# separar el encabezado no lo hace desaparecer, lo convierte en una unidad
# propia — y una unidad que no termina en punto sigue sin terminar en punto
# cuando le toca caer al final de un chunk. El caso de referencia del brief,
# "…no access. 1.3 Frontier AI Research Chapter 1: Research and Development",
# sigue contando como corte_real después del fix. La métrica solo baja si el
# encabezado se BORRA (perder contenido recuperable) o si el empaquetador
# evita cerrar chunk sobre una unidad corta que no cierra oración — y eso
# último vive en chunk.empaquetar(), no aquí.
MAX_PALABRAS_NO_PROSA = 12

# Numeración de sección al inicio: "1.1 Publications", "2 Antecedentes",
# "3.4.1 Método". Exige texto detrás: "2010" solo es un dato, no un título.
_RE_SECCION_NUMERADA = re.compile(r"^\d+(?:\.\d+)*[\.\)]?\s+\S")

# Etiqueta de dato de gráfica: empieza por número/porcentaje y una coma.
# "0.05, Clinical trial", "1.46%, Other", "12,052, Other".
_RE_ETIQUETA_DATO = re.compile(r"^[\d.,%$€]+\s*,\s*\S")

# Terminadores de oración a efectos de "esta línea cerró". Mismo criterio que
# scripts/medir_terminadores.py, para que lo que se arregla y lo que se mide
# no puedan divergir.
_TERMINADORES_LINEA = tuple(".!?;:\"')]}»…”’")


def _es_titulo(linea: str) -> bool:
    """
    Formato título: mayoría de palabras alfabéticas con inicial mayúscula.

    Deja fuera los line-wraps de prosa, que es lo importante: "Between 2010 and
    2022, the total number of AI" tiene una sola inicial mayúscula de seis
    palabras y no pasa. "Chapter Highlights" o "Foundation Models", sí.
    """
    palabras = [p for p in linea.split() if p[:1].isalpha()]
    if len(palabras) < 2:
        return False
    mayusculas = sum(1 for p in palabras if p[:1].isupper())
    return mayusculas / len(palabras) >= 0.6


def _es_linea_no_prosa(linea: str) -> bool:
    """Forma de encabezado de sección o de etiqueta de gráfica."""
    linea = linea.strip()
    if not linea or len(linea.split()) > MAX_PALABRAS_NO_PROSA:
        return False
    if linea.endswith(_TERMINADORES_LINEA):
        # Ya cierra: no se pega a lo que sigue, no hay nada que separar.
        return False
    return bool(
        _RE_SECCION_NUMERADA.match(linea)
        or _RE_ETIQUETA_DATO.match(linea)
        or _es_titulo(linea)
        or (linea.isupper() and any(c.isalpha() for c in linea))
    )


def aislar_lineas_no_prosa(paginas: list[str]) -> tuple[list[str], dict]:
    """
    Rodea de líneas en blanco las líneas que no son prosa, sin borrar ninguna.

    Solo actúa cuando el contexto confirma que separar no rompe nada:
      * la línea anterior cerró oración (o la línea abre la página), y
      * la siguiente no empieza en minúscula, que sería continuación de esta.

    Returns:
        (paginas, diagnostico) con cuántas líneas se aislaron y una muestra,
        para que ingest_data.py lo registre: una separación silenciosa es tan
        difícil de auditar como un borrado silencioso.
    """
    salida, aisladas, muestra = [], 0, []

    for pagina in paginas:
        lineas = [l.strip() for l in pagina.splitlines() if l.strip()]
        if not lineas:
            salida.append("")
            continue

        piezas = []
        for i, linea in enumerate(lineas):
            previa = lineas[i - 1] if i else None
            siguiente = lineas[i + 1] if i + 1 < len(lineas) else None

            cierra_previa = previa is None or previa.endswith(_TERMINADORES_LINEA)
            # Si la siguiente arranca en minúscula, esta línea es el principio
            # de una oración partida por el ancho de página, no un encabezado.
            continua_siguiente = bool(siguiente) and siguiente[:1].islower()

            if cierra_previa and not continua_siguiente and _es_linea_no_prosa(linea):
                aisladas += 1
                if len(muestra) < 10:
                    muestra.append(linea)
                piezas.append("\n" + linea + "\n")   # el "\n\n" lo da el join
            else:
                piezas.append(linea)

        salida.append("\n".join(piezas))

    return salida, {"lineas_aisladas": aisladas, "muestra_aisladas": muestra}


# --- Oraciones que cruzan de página -----------------------------------------
#
# APAGADO POR DEFECTO, igual que el bloque anterior y por el mismo motivo:
# sobre la muestra dispara solo 25 veces en ~800 páginas y no mueve la métrica
# (433 -> 433 corte_real). La hipótesis era razonable y resultó marginal: casi
# ninguna página termina a media oración con la siguiente empezando en
# minúscula, porque entre las dos mitades suele quedar el encabezado de la
# página siguiente.
#
# Medido sobre la muestra: la mayoría de los chunks que cortan a media frase no
# cortan en un encabezado, cortan en el SALTO DE PÁGINA. Las páginas se unían
# siempre con "\n\n", chunk.py parte párrafo justo ahí (`re.split(r"\n\s*\n")`)
# y una oración que sigue en la página siguiente queda partida en dos unidades.
# El chunker no puede recomponerla: nunca junta lo que recibe separado.
#
# La regla es la misma que chunk._unir_line_wraps() aplica DENTRO del párrafo,
# subida un nivel: se une solo si la página anterior no cerró oración y la
# siguiente arranca en minúscula. Si arranca en mayúscula o dígito es un
# título, una nueva oración o un ítem de índice, y el salto se respeta.
def unir_paginas(paginas: list[str]) -> tuple[str, dict]:
    """Une las páginas, sin separación de párrafo cuando la oración continúa."""
    partes, unidas = [], 0

    for pagina in paginas:
        if not pagina.strip():
            continue
        if not partes:
            partes.append(pagina)
            continue

        anterior = partes[-1].rstrip()
        primera = pagina.lstrip()[:1]
        continua = (anterior
                    and not anterior.endswith(_TERMINADORES_LINEA)
                    and primera.islower())
        if continua:
            unidas += 1
            partes[-1] = partes[-1].rstrip() + "\n" + pagina.lstrip()
        else:
            partes.append(pagina)

    return "\n\n".join(partes), {"paginas_unidas": unidas}


# --- Puntuación desplazada en líneas árabes (RTL) ---------------------------
#
# APAGADO POR DEFECTO (`_aplicar_limpieza(rtl=False)`), a la espera de medición.
#
# QUÉ PASA DE VERDAD, que no es lo que parecía.
# El diagnóstico inicial decía que PyMuPDF emite los runs árabes "en orden
# visual invertido". Comprobado punto de código a punto de código, es más
# estrecho: las LETRAS y el ORDEN DE PALABRAS son correctos. Lo único que se
# desplaza es la puntuación de cierre de la línea.
#
# En una línea RTL el punto final se dibuja en el extremo IZQUIERDO. PyMuPDF
# recorre la página de izquierda a derecha, así que lo emite primero y queda
# pegado al inicio:
#
#     emitido : ".جميع الحقوق محفوظة"      <- el punto abre la línea
#     correcto: "جميع الحقوق محفوظة."      <- el punto cierra la línea
#                (= "todos los derechos reservados.")
#
# Consecuencia: la línea nunca termina en terminador, así que todo chunk que
# cierre ahí cuenta como `corte_real`.
#
# AVISO PARA QUIEN VERIFIQUE ESTO A MANO: la terminal aplica bidi al pintar
# árabe, de modo que un texto lógicamente invertido SE VE correcto y viceversa.
# Cualquier comprobación visual en consola miente. Verificar con
# unicodedata.name() sobre los puntos de código, nunca a ojo.
_RE_ARABE = re.compile(r"[؀-ۿݐ-ݿﭐ-﷿ﹰ-﻿]")

# Solo puntuación que CIERRA ORACIÓN. Medido: incluir coma (،), punto y coma
# (؛) y dos puntos empeoraba las cosas. En F2-UNOOSA-025, 100 de 135 movimientos
# eran `؛`: mover ese signo no cierra ninguna oración, pero sí cambia dónde
# segmenta pysbd y baraja las fronteras de chunk. El documento empeoraba 2,6pp
# por puro churn. Un signo que no cierra oración no arregla un corte de oración.
_PUNTUACION_CIERRE_RTL = ".!?…؟"

# Mínimo de caracteres árabes para considerar que la línea es RTL. Por debajo
# es una cita suelta dentro de un texto latino y no se toca.
MIN_ARABE_LINEA = 8

# Fracción mínima de letras que deben ser árabes. Evita actuar sobre líneas
# mixtas donde el latín manda y la dirección del párrafo es LTR.
MIN_FRACCION_ARABE = 0.5


def _es_linea_rtl(linea: str) -> bool:
    """La línea es predominantemente árabe, no una cita suelta."""
    arabes = len(_RE_ARABE.findall(linea))
    if arabes < MIN_ARABE_LINEA:
        return False
    letras = sum(1 for c in linea if c.isalpha())
    return letras > 0 and arabes / letras >= MIN_FRACCION_ARABE


def corregir_puntuacion_rtl(texto: str) -> tuple[str, dict]:
    """
    Devuelve al final la puntuación que quedó al inicio de las líneas árabes.

    Conservador a propósito: solo actúa si la línea es mayoritariamente árabe,
    empieza por puntuación de cierre y NO termina ya en una. Si termina en
    puntuación, la del inicio pertenece a otra cosa y moverla inventaría texto.
    """
    lineas = texto.split("\n")
    salida, movidas = [], 0

    for linea in lineas:
        despojada = linea.strip()
        if (not despojada
                or despojada[0] not in _PUNTUACION_CIERRE_RTL
                or despojada[-1] in _PUNTUACION_CIERRE_RTL
                or not _es_linea_rtl(despojada)):
            salida.append(linea)
            continue

        corte = 0
        while corte < len(despojada) and despojada[corte] in _PUNTUACION_CIERRE_RTL:
            corte += 1
        prefijo, resto = despojada[:corte], despojada[corte:].lstrip()
        if not resto:
            salida.append(linea)
            continue

        salida.append(resto + prefijo)
        movidas += 1

    return "\n".join(salida), {"rtl_puntuacion_movida": movidas}


# --- cmap roto: texto desplazado en el punto de codigo ----------------------
#
# APAGADO POR DEFECTO (`_aplicar_limpieza(cmap=False)`).
#
# CORRIJO MI PROPIO DIAGNOSTICO ANTERIOR. El informe decia que en F3-CEOBS-030
# "las fronteras de palabra se perdieron en la extraccion" y que por eso solo
# quedaba vaciar el documento. Es falso para el grueso del texto: el CUERPO
# conserva los espacios y se descifra limpio. Lo que se pierde son los espacios
# de los titulares y del indice, que ya venian pegados en el PDF.
#
# Tampoco es un Cesar sobre a-z, como escribi: es un desplazamiento uniforme
# del PUNTO DE CODIGO, que cruza la frontera entre letras, digitos y simbolos.
#
#     "8LI 1MREQEXE" +28 -> "The Minamata"      T=84, 8=56  -> 28 exactos
#     "/LVWRI"       +29 -> "Listof"            L=76, /=47  -> 29 exactos
#
# Son subconjuntos de fuente distintos, cada uno con su cmap roto a su manera,
# no un cifrado unico.
OFFSETS_CMAP = (28, 29, -1)

# Fraccion de vocales del ingles sano. Medida sobre 994 documentos en ingles del
# corpus: p05=0,376  mediana=0,391  p95=0,406. Se usa como juez porque NO
# depende de los espacios: funciona igual en "Listof" que en "List of". El juez
# de palabras funcionales que probe primero daba cero en el texto pegado y
# empataba con "no tocar", asi que ganaba no tocar.
VOCALES_INGLES = 0.39
TOLERANCIA_VOCALES = 0.09

# Minimo de letras latinas por linea. Por debajo, la fraccion de vocales es
# ruido estadistico y un acierto por azar reescribiria texto sano.
MIN_LETRAS_CMAP = 12

# Puerta a nivel de DOCUMENTO. El descifrado solo se intenta si el documento
# entero puntua como corrupto. Sin esta puerta, la funcion podria reescribir
# texto legitimo en algun documento con mucha tabla o sigla.
#
# El umbral cae en mitad de un hueco enorme. Medido sobre los 1.721 documentos
# del corpus con >=500 letras latinas:
#
#     F3-CEOBS-030   0,1970   <- el unico corrupto
#     F2-SWF-130     0,3159   <- el siguiente mas bajo
#
# Con 0,25 la puerta selecciona 1 documento de 1.721 y sobra margen por los dos
# lados. Ojo: un umbral calibrado con datos de PARRAFO no vale aqui — los
# parrafos sanos bajan mucho mas que los documentos enteros, y con 0,15 (la
# cifra por parrafo) la puerta no disparaba ni en CEOBS-030.
MAX_VOCALES_DOC_CORRUPTO = 0.25

_VOCALES = frozenset("aeiouAEIOU")


def _fraccion_vocales(texto: str) -> float | None:
    """Vocales sobre letras latinas. None si no hay letras latinas que medir."""
    letras = [c for c in texto if "a" <= c.lower() <= "z"]
    if not letras:
        return None
    return sum(1 for c in letras if c in _VOCALES) / len(letras)


def _desplazar_codigo(texto: str, k: int) -> str:
    """Desplaza el punto de codigo de los ASCII imprimibles, nada mas."""
    return "".join(
        chr(ord(c) + k) if 32 < ord(c) < 127 else c
        for c in texto
    )


def corregir_cmap_desplazado(texto: str) -> tuple[str, dict]:
    """
    Deshace el desplazamiento de punto de codigo de un PDF con el cmap roto.

    Conservador en tres niveles: el documento entero tiene que puntuar como
    corrupto, la linea tiene que tener letras suficientes, y el descifrado tiene
    que caer dentro de la tolerancia del ingles sano. Si algo no cuadra, la
    linea se devuelve intacta: es preferible dejar ruido que reescribir texto
    bueno.
    """
    global_voc = _fraccion_vocales(texto)
    if global_voc is None or global_voc >= MAX_VOCALES_DOC_CORRUPTO:
        return texto, {}

    salida, arregladas = [], 0
    por_offset: dict[int, int] = {}

    for linea in texto.split("\n"):
        voc = _fraccion_vocales(linea)
        letras = sum(1 for c in linea if "a" <= c.lower() <= "z")
        if (voc is None
                or letras < MIN_LETRAS_CMAP
                or abs(voc - VOCALES_INGLES) <= TOLERANCIA_VOCALES):
            salida.append(linea)          # vacia, corta, o ya parece sana
            continue

        mejor_k, mejor_dist = None, None
        for k in OFFSETS_CMAP:            # orden fijo: desempate determinista
            cand = _fraccion_vocales(_desplazar_codigo(linea, k))
            if cand is None:
                continue
            dist = abs(cand - VOCALES_INGLES)
            if mejor_dist is None or dist < mejor_dist:
                mejor_k, mejor_dist = k, dist

        if mejor_k is None or mejor_dist > TOLERANCIA_VOCALES:
            salida.append(linea)
            continue

        salida.append(_desplazar_codigo(linea, mejor_k))
        arregladas += 1
        por_offset[mejor_k] = por_offset.get(mejor_k, 0) + 1

    if not arregladas:
        return texto, {}

    return "\n".join(salida), {
        "cmap_lineas_corregidas": arregladas,
        "cmap_offsets": {str(k): por_offset[k] for k in sorted(por_offset)},
    }


def vaciar_cmap_roto(texto: str) -> tuple[str, dict]:
    """
    Opcion (b): descartar el texto ilegible en vez de descifrarlo.

    Se conserva por si el descifrado no pasa el gate. Vacia las lineas que
    puntuan como corruptas y no toca las demas. No inventa contenido: solo
    borra lo que ya era ruido para el indice vectorial.
    """
    global_voc = _fraccion_vocales(texto)
    if global_voc is None or global_voc >= MAX_VOCALES_DOC_CORRUPTO:
        return texto, {}

    salida, vaciadas = [], 0
    for linea in texto.split("\n"):
        voc = _fraccion_vocales(linea)
        letras = sum(1 for c in linea if "a" <= c.lower() <= "z")
        if (voc is not None
                and letras >= MIN_LETRAS_CMAP
                and abs(voc - VOCALES_INGLES) > TOLERANCIA_VOCALES):
            vaciadas += 1
            continue
        salida.append(linea)

    if not vaciadas:
        return texto, {}
    return "\n".join(salida), {"cmap_lineas_vaciadas": vaciadas}


# Fixes de limpieza APROBADOS para la corrida de produccion, en un solo sitio
# para que las tres llamadas de extraer_pdf no se desincronicen. Los parametros
# de `_aplicar_limpieza` siguen apagados por defecto a proposito: asi los A/B de
# scripts/ miden contra la base real y no contra la configuracion aprobada.
#
#   rtl   — puntuacion RTL desplazada. Medido: -8 corte_real en la muestra,
#           3 documentos afectados, 17 controles a +0,0pp.
#   cmap  — cmap roto de F3-CEOBS-030. Medido: legibilidad 0,505 -> 0,960,
#           87,6% de los caracteres recuperados, 1 documento de 1.721.
#
# `aislar` y `unir` NO entran: medidos y no pagan.
LIMPIEZA_APROBADA = {"rtl": True, "cmap": "descifrar"}


def _aplicar_limpieza(paginas: list[str], aislar: bool = False,
                      unir: bool = False, rtl: bool = False,
                      cmap: str | None = None) -> tuple[str, dict]:
    """
    Limpia el boilerplate y aplica la salvaguarda del 40%.

    Devuelve (texto, extra_parcial). Si la limpieza se pasa de la raya, se
    devuelve el texto ORIGINAL y el diagnóstico dice por qué, para que el
    documento se revise a mano en vez de quedarse vaciado.

    `aislar` y `unir` existen para el A/B de scripts/ab_extraccion.py: permiten
    medir cada fix por separado sin tocar el módulo.
    """
    original = "\n\n".join(paginas)
    limpias, diag = limpiar_boilerplate_paginas(paginas)
    if aislar:
        limpias, diag_aislado = aislar_lineas_no_prosa(limpias)
        diag.update(diag_aislado)
    if unir:
        limpio, diag_unido = unir_paginas(limpias)
        diag.update(diag_unido)
    else:
        limpio = "\n\n".join(p for p in limpias if p.strip())

    # Va al final, sobre el texto ya montado: opera línea a línea y no depende
    # de la separación por páginas.
    if rtl:
        limpio, diag_rtl = corregir_puntuacion_rtl(limpio)
        diag.update(diag_rtl)

    # `cmap` es "descifrar" (a) o "vaciar" (b); None deja el texto como esta.
    if cmap == "descifrar":
        limpio, diag_cmap = corregir_cmap_desplazado(limpio)
        diag.update(diag_cmap)
    elif cmap == "vaciar":
        limpio, diag_cmap = vaciar_cmap_roto(limpio)
        diag.update(diag_cmap)
    elif cmap is not None:
        raise ValueError(f"cmap debe ser 'descifrar', 'vaciar' o None: {cmap!r}")

    palabras_antes = len(original.split())
    palabras_despues = len(limpio.split())
    if not palabras_antes:
        return original, {}

    fraccion = 1 - (palabras_despues / palabras_antes)
    extra = {
        "limpieza_lineas": diag.get("lineas_eliminadas", 0),
        "limpieza_fraccion_palabras": round(fraccion, 4),
    }
    if diag.get("lineas_aisladas"):
        extra["limpieza_aisladas"] = diag["lineas_aisladas"]
    if diag.get("paginas_unidas"):
        extra["limpieza_paginas_unidas"] = diag["paginas_unidas"]
    if diag.get("rtl_puntuacion_movida"):
        extra["rtl_puntuacion_movida"] = diag["rtl_puntuacion_movida"]
    if diag.get("cmap_lineas_corregidas"):
        extra["cmap_lineas_corregidas"] = diag["cmap_lineas_corregidas"]
        extra["cmap_offsets"] = diag["cmap_offsets"]
    if diag.get("cmap_lineas_vaciadas"):
        extra["cmap_lineas_vaciadas"] = diag["cmap_lineas_vaciadas"]
    if diag.get("n_patrones"):
        extra["limpieza_patrones"] = diag["patrones_detectados"]

    if fraccion > MAX_FRACCION_ELIMINABLE:
        extra["limpieza_omitida"] = True
        extra["limpieza_motivo"] = (
            f"habría eliminado el {fraccion * 100:.1f}% de las palabras "
            f"(máximo {MAX_FRACCION_ELIMINABLE * 100:.0f}%): requiere revisión manual"
        )
        return original, extra

    return limpio, extra


# --- OCR de PDF escaneados --------------------------------------------------
#
# DPI de rasterizacion. FIJO Y EXPLICITO a proposito: junto con la version de
# Tesseract es lo que hace el OCR reproducible. Si se deja el valor por defecto
# de PyMuPDF, una actualizacion de la libreria cambia la resolucion, cambia el
# texto reconocido y rompe el criterio 9 (dos corridas -> mismo sha256).
# Verificado a 300 DPI con Tesseract 5.5.3: 5 informes, dos procesos separados,
# 5/5 sha256 identicos. Ver el pin en requirements.txt.
DPI_OCR_PDF = 300
IDIOMAS_OCR_PDF = "spa+eng"

# Umbral de disparo del OCR, en CARACTERES por pagina.
#
# Se mide en caracteres y no en palabras porque el corpus tiene PDF en chino,
# donde no hay espacios: por palabras, un informe chino denso puntua 27 pal/pag
# y parece vacio. Por caracteres puntua 465.
#
# El valor 50 no es arbitrario, sale de la distribucion real de los 759 PDF:
#     51 documentos entre 0.0 y 3.6 char/pag   (escaneados sin capa de texto)
#     --- hueco de 100 char/pag, CERO documentos ---
#     el siguiente esta en 103.2
# Es decir, el corte vive en una banda vacia de 3.6 a 103.2, un factor de 29x.
# Que una version distinta de PyMuPDF mueva un documento de un lado al otro
# exigiria que su extraccion cambiara en un orden de magnitud.
MIN_CARACTERES_PDF_POR_PAGINA = 50


def _ocr_pdf(ruta: Path, paginas_max: int | None = None) -> list[str]:
    """
    Rasteriza cada pagina a DPI fijo y le pasa Tesseract.

    Devuelve la lista POR PAGINA, no el texto unido: los escaneados tambien
    llevan encabezado repetido, y sin la separacion por pagina no se puede
    detectar. Une el llamador, despues de limpiar.
    """
    import io

    import fitz
    import pytesseract
    from PIL import Image

    doc = fitz.open(ruta.as_posix())
    try:
        partes = []
        for i, pagina in enumerate(doc):
            if paginas_max is not None and i >= paginas_max:
                break
            pix = pagina.get_pixmap(dpi=DPI_OCR_PDF)
            imagen = Image.open(io.BytesIO(pix.tobytes("png")))
            partes.append(pytesseract.image_to_string(imagen, lang=IDIOMAS_OCR_PDF))
    finally:
        doc.close()
    return partes


def extraer_pdf(ruta: Path, usar_ocr: bool = True) -> tuple[str, str | None, dict]:
    """Extrae texto de un PDF intentando usar PyMuPDF (fitz) o PyPDF2.

    Si el PDF no trae capa de texto (esta escaneado), cae a OCR: son 51 de los
    759 del corpus, 45 de ellos informes de Alertas Tempranas. Sin OCR esos
    documentos tienen DOC_ID y metadata correctos pero nada que recuperar.

    Si no hay ninguna dependencia, lanza ValueError informativo para que
    el llamador lo registre.

    El `extra` que devuelve NO es decorativo: lleva el motor que acabó
    extrayendo (`motor_pdf`), el número de páginas y, si hubo que recurrir a
    PyPDF2, el error de PyMuPDF que lo provocó (`aviso_pdf`). Antes ese
    fallback era silencioso: un PDF que PyMuPDF abría a medias caía a PyPDF2
    sin dejar rastro, y un texto truncado es indistinguible de uno completo
    mirando solo el JSONL. `ingest_data.py` vuelca estos avisos al log.
    """
    # Intentar PyMuPDF primero (mejor calidad y preserva orden)
    motivo_fallback = None
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(ruta.as_posix())
        try:
            partes = []
            for pagina in doc:
                texto = pagina.get_text("text")
                if texto:
                    partes.append(texto)
            n_paginas = doc.page_count
        finally:
            doc.close()
        texto, extra_limpieza = _aplicar_limpieza(partes, **LIMPIEZA_APROBADA)
        extra = {"motor_pdf": "pymupdf", "paginas": n_paginas, **extra_limpieza}

        # ¿Trae capa de texto? Si no, está escaneado y hay que rasterizar.
        if usar_ocr and n_paginas and (
            len(texto) / n_paginas < MIN_CARACTERES_PDF_POR_PAGINA
        ):
            caracteres_previos = len(texto)
            paginas_ocr = _ocr_pdf(ruta)
            texto_ocr, extra_limpieza_ocr = _aplicar_limpieza(
                paginas_ocr, **LIMPIEZA_APROBADA)
            # Solo se acepta el OCR si aporta más de lo que ya había: si
            # Tesseract devuelve menos, el original imperfecto es mejor que
            # una página en blanco.
            if len(texto_ocr) > caracteres_previos:
                texto = texto_ocr
                extra = {
                    "motor_pdf": "ocr",
                    "motor_pdf_previo": "pymupdf",
                    "paginas": n_paginas,
                    "ocr_dpi": DPI_OCR_PDF,
                    "ocr_idiomas": IDIOMAS_OCR_PDF,
                    "caracteres_capa_texto": caracteres_previos,
                    **extra_limpieza_ocr,
                }
        return texto, None, extra
    except ImportError as e:
        motivo_fallback = f"PyMuPDF no disponible: {e}"
    except Exception as e:  # noqa: BLE001
        motivo_fallback = f"PyMuPDF falló: {type(e).__name__}: {e}"

    # Intentar PyPDF2 como alternativa. Se llega aquí solo si PyMuPDF falló,
    # y el motivo viaja en `extra` para que quede en el log de fallos.
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(ruta)
        partes = []
        paginas_fallidas = 0
        for p in reader.pages:
            try:
                t = p.extract_text() or ""
            except Exception:  # noqa: BLE001
                t = ""
                paginas_fallidas += 1
            if t:
                partes.append(t)
        texto, extra_limpieza = _aplicar_limpieza(partes, **LIMPIEZA_APROBADA)
        extra = {
            "motor_pdf": "pypdf2",
            "paginas": len(reader.pages),
            "aviso_pdf": motivo_fallback,
            **extra_limpieza,
        }
        if paginas_fallidas:
            extra["paginas_fallidas"] = paginas_fallidas
        return texto, None, extra
    except Exception as e:  # noqa: BLE001
        raise ValueError(
            "El extractor de PDF requiere 'PyMuPDF' (fitz) o 'PyPDF2' instalado. "
            f"PyMuPDF: {motivo_fallback}. PyPDF2: {type(e).__name__}: {e}"
        )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def extraer_csv(ruta: Path) -> tuple[str, str | None, dict]:
    """
    Serializa un CSV como "columna: valor | columna: valor", una fila por línea.

    Mismo contrato de serialización que extraer_xlsx(): son el mismo tipo de
    dato (registros tabulares) y el chunker los trata igual — la unidad atómica
    es la línea. Las celdas vacías se omiten en vez de emitir "columna: ",
    que solo mete ruido en el índice.

    Cada valor pasa por aplanar_valor(): una celda CSV puede contener saltos de
    línea dentro de comillas, y sin aplanar la fila se parte en varias líneas
    del JSONL y deja de ser un registro atómico. Es exactamente el bug que ya
    se corrigió en xlsx y pbf.

    Detecta el delimitador con csv.Sniffer (hay CSV separados por ';' en el
    corpus) y cae a ',' si no logra decidir.
    """
    # utf-8-sig: varios CSV del corpus traen BOM y sin esto la primera columna
    # de la cabecera se llama "﻿nombre" y no casa con nada.
    texto_bruto = ruta.read_text(encoding="utf-8-sig", errors="replace")
    if not texto_bruto.strip():
        return "", ruta.stem, {"filas": 0}

    muestra = texto_bruto[:8192]
    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        delimitador = dialecto.delimiter
    except csv.Error:
        delimitador = ","

    lector = csv.reader(texto_bruto.splitlines(), delimiter=delimitador)

    cabecera = None
    lineas = []
    for fila in lector:
        if not fila or all(not str(c).strip() for c in fila):
            continue

        if cabecera is None:
            cabecera = [formatear_valor(c).strip() for c in fila]
            continue

        partes = []
        for col, valor in zip(cabecera, fila):
            if not col:
                continue
            texto_valor = aplanar_valor(formatear_valor(valor))
            if texto_valor:
                partes.append(f"{col}: {texto_valor}")
        if partes:
            lineas.append(" | ".join(partes))

    extra = {"filas": len(lineas)}
    if cabecera:
        extra["columnas"] = [c for c in cabecera if c]
    if delimitador != ",":
        extra["delimitador"] = delimitador

    return "\n".join(lineas), ruta.stem, extra



# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def extraer_xlsx(ruta: Path) -> tuple[str, str | None, dict]:
    """
    Serializa un XLSX como "columna: valor | columna: valor", una fila por línea.

    El nombre de la hoja se incluye como contexto: es información semántica
    real (p.ej. "Private investment" vs "Publications by country").

    Nota: en read_only=True openpyxl no siempre calcula ws.max_row, así que
    se itera directamente en vez de confiar en las dimensiones.
    """
    import openpyxl

    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    bloques = []
    hojas_con_datos = []

    try:
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            filas = ws.iter_rows(values_only=True)

            cabecera = None
            lineas_hoja = []

            for fila in filas:
                if fila is None:
                    continue
                celdas = [c for c in fila if c is not None]
                if not celdas:
                    continue

                if cabecera is None:
                    cabecera = [formatear_valor(c) if c is not None else ""
                                for c in fila]
                    continue

                partes = []
                for col, valor in zip(cabecera, fila):
                    if valor is None or not col:
                        continue
                    # Una celda puede traer saltos de línea (texto envuelto o
                    # listas). Sin aplanar, la fila se parte en varias líneas
                    # del JSONL y deja de ser un registro atómico.
                    texto_valor = aplanar_valor(formatear_valor(valor))
                    if texto_valor:
                        partes.append(f"{col}: {texto_valor}")
                if partes:
                    lineas_hoja.append(" | ".join(partes))

            if lineas_hoja:
                hojas_con_datos.append(nombre_hoja)
                bloques.append(f"hoja: {nombre_hoja}\n" + "\n".join(lineas_hoja))
    finally:
        wb.close()

    # Las hojas van también en `extra` para que el chunker pueda anteponerlas
    # como cabecera de CADA chunk. En el texto, "hoja: X" solo aparece en la
    # primera línea, así que sin esto únicamente el chunk c000 tendría contexto.
    # Se deja además en el texto por seguridad: si C no llega a implementar la
    # cabecera, el contexto no se pierde del todo.
    extra = {}
    if hojas_con_datos:
        extra["hojas"] = hojas_con_datos
        # Nombres genéricos de openpyxl: no aportan señal semántica y conviene
        # que C sepa cuáles NO merece la pena anteponer.
        utiles = [h for h in hojas_con_datos
                  if not re.fullmatch(r"(?i)(sheet|hoja|feuille|planilha)\s*\d*", h.strip())]
        extra["hojas_utiles"] = utiles

    return "\n\n".join(bloques), ruta.stem, extra


# ---------------------------------------------------------------------------
# Imágenes (OCR)
# ---------------------------------------------------------------------------

def extraer_imagen(ruta: Path, idiomas: str = "spa+eng+por") -> tuple[str, str | None, dict]:
    """
    Aplica OCR multilingüe a una imagen.

    Requiere el binario de Tesseract instalado en el sistema (no basta con
    pip install pytesseract) con los paquetes de idioma spa, eng y por.

    El filtrado de resultados pobres NO se hace aquí: lo decide ingest_data.py
    con text_is_usable(), para que el criterio quede en un solo sitio.
    """
    import pytesseract
    from PIL import Image

    # AVIF necesita un plugin externo; si no está, Pillow lanza
    # "Unsupported image format/type". Se intenta cargar en silencio.
    try:
        import pillow_avif  # noqa: F401
    except ImportError:
        pass

    with Image.open(ruta) as img:
        ancho, alto = img.size
        # pytesseract mantiene una LISTA BLANCA de formatos (BMP, GIF, JPEG,
        # JPEG2000, PBM, PGM, PNG, PPM, TIFF, WEBP) y AVIF no está en ella:
        # aunque Pillow abra el archivo con pillow_avif, prepare() lanza
        # TypeError('Unsupported image format/type') al ver image.format='AVIF'.
        # convert() devuelve una imagen con format=None, y pytesseract trata
        # ese caso como PNG. Además normaliza el modo (P, LA, CMYK) y descarta
        # el canal alfa, que es otra fuente de fallos.
        texto = pytesseract.image_to_string(img.convert("RGB"), lang=idiomas)

    extra = {"ancho_px": ancho, "alto_px": alto}
    return texto, None, extra


# ---------------------------------------------------------------------------
# PBF (Mapbox Vector Tiles)
# ---------------------------------------------------------------------------

def parsear_zoom_x_y(ruta_relativa: str) -> dict:
    """
    Extrae zoom/x/y de una ruta tipo "tiles/3/2/AMAZONUW_4.pbf".

    El nombre del archivo lleva el prefijo del observatorio, así que el
    'y' hay que sacarlo del nombre quitando ese prefijo.
    """
    partes = ruta_relativa.replace("\\", "/").split("/")
    if "tiles" not in partes:
        return {}

    idx = partes.index("tiles")
    siguientes = partes[idx + 1:]
    if len(siguientes) < 3:
        return {}

    zoom, x = siguientes[0], siguientes[1]
    nombre = Path(siguientes[2]).stem
    y = nombre.split("_")[-1]  # "AMAZONUW_4" -> "4"

    if zoom.isdigit() and x.isdigit() and y.isdigit():
        return {"zoom": int(zoom), "x": int(x), "y": int(y)}
    return {}


def extraer_pbf(
    ruta: Path,
    ruta_relativa: str,
    features_asignados: dict | None = None,
) -> tuple[str, str | None, dict]:
    """
    Decodifica un Mapbox Vector Tile y serializa los atributos de sus features.

    Se priorizan los campos au_popup_window_* : son texto ya redactado para
    lectura humana, mucho mejor que la serialización cruda de atributos.

    La geometría y los códigos internos se descartan: no aportan nada
    semánticamente a un encoder de texto.

    Deduplicación: si se pasa features_asignados ({fid: props}), solo se
    emiten esos features, y se usan LOS PROPS DEL DICCIONARIO, no los del
    tile. Así el feature se indexa en el tile más granular pero conserva
    la versión más rica de sus atributos (ver src/dedup_pbf.py). El corpus tiene 88% de redundancia entre
    niveles de zoom (11.906 features, 1.409 únicos); sin esto, una consulta
    recuperaría hasta 26 fragmentos idénticos que llenarían el top-10.
    El conjunto lo calcula src/dedup_pbf.py.
    """
    import mapbox_vector_tile

    from dedup_pbf import CLAVE_RESUMEN

    datos = ruta.read_bytes()
    tile = mapbox_vector_tile.decode(datos)

    bloques = []
    total_features = 0
    features_indexados = 0

    # Tile de zoom bajo cuyos features ya se indexan en tiles más granulares:
    # emite un resumen de cobertura en vez de quedarse vacío. Un documento sin
    # texto no es recuperable, y cada tile es un `fuente` del ground truth.
    if features_asignados and CLAVE_RESUMEN in features_asignados:
        total_features = sum(
            len(capa.get("features", [])) for capa in tile.values()
        )
        extra = parsear_zoom_x_y(ruta_relativa)
        extra["num_features_tile"] = total_features
        extra["num_features_indexados"] = 0
        extra["capas"] = list(tile.keys())
        extra["es_resumen"] = True
        return features_asignados[CLAVE_RESUMEN], None, extra

    for nombre_capa, capa in tile.items():
        features = capa.get("features", [])
        total_features += len(features)
        lineas = []

        for feature in features:
            props = feature.get("properties", {}) or {}

            # Deduplicación: este feature se indexa en otro tile.
            if features_asignados is not None:
                fid = props.get("fid")
                if fid is None or fid not in features_asignados:
                    continue
                # Usar la versión más rica encontrada en todo el dataset.
                props = features_asignados[fid]

            # 1. Texto ya redactado, si existe.
            # Los popups vienen redactados para lectura humana y suelen traer
            # listas con viñetas en varias líneas. Es el origen del 59.7% de
            # líneas huérfanas detectadas en el JSONL: hay que aplanarlos.
            popups = [
                aplanar_valor(formatear_valor(props[k]))
                for k in ATRIBUTOS_PBF_PRIORITARIOS
                if props.get(k) not in (None, "")
            ]
            popups = [p for p in popups if p]

            # Deduplicar: las variantes es/pt/en del popup suelen traer el
            # MISMO texto (los nombres de grupos armados son propios y no se
            # traducen). Sin esto se indexa triplicado — "Los Lobos Los Lobos
            # Los Lobos" — , que infla el chunk sin añadir señal.
            vistos = set()
            unicos = []
            for p in popups:
                if p not in vistos:
                    vistos.add(p)
                    unicos.append(p)

            # Prefijo "popup:" para que TODA línea de registro tenga la forma
            # "campo: valor". Sin él esta rama emitía líneas sin prefijo,
            # indistinguibles de un fragmento huérfano al validar.
            if unicos:
                lineas.append("popup: " + "; ".join(unicos))
                features_indexados += 1
                continue

            # 2. Si no hay popup, serializar atributos semánticos.
            partes = []
            for clave in ATRIBUTOS_PBF_SEMANTICOS:
                if clave in ATRIBUTOS_PBF_DESCARTADOS:
                    continue
                valor = props.get(clave)
                if valor in (None, ""):
                    continue
                partes.append(f"{clave}: {aplanar_valor(formatear_valor(valor))}")
            if partes:
                lineas.append(" | ".join(partes))
                features_indexados += 1

        if lineas:
            bloques.append(f"capa: {nombre_capa}\n" + "\n".join(lineas))

    extra = parsear_zoom_x_y(ruta_relativa)
    extra["num_features_tile"] = total_features
    # Se cuentan features, no líneas. Antes se contaban líneas y los popups
    # multilínea inflaban la cifra: un feature de 4 líneas contaba como 4.
    extra["num_features_indexados"] = features_indexados
    extra["capas"] = list(tile.keys())

    return "\n\n".join(bloques), None, extra