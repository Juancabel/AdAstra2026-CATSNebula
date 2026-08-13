# -*- coding: utf-8 -*-
"""
indice_oficial.py — Carga del inventario de ADL (Indice_Datos_Codefest.xlsx).

POR QUÉ ESTE MÓDULO
El índice tiene 1.826 filas pero solo 1.699 nombres de archivo distintos: 127
archivos comparten nombre con otro. Indexar por `Nombre estandarizado` los
pisa entre sí y deja el inventario en 1.699, que es lo que reportaba la
ingesta. Con ese total, la verificación de completitud daría por completa una
ingesta a la que le faltan 127 archivos.

Los tres grupos que colisionan por nombre:
  - 72 teselas PBF   (tiles/<z>/<x>/AMAZONUW_<y>.pbf se repite entre zooms)
  - 112 archivos CSET (mismo informe en pdfs/Reports y pdfs/Translation)
  -   2 archivos ESA  (mismo PDF archivado en dos carpetas)

La clave correcta es `Carpeta + "/" + Nombre estandarizado`, que da 1.826
valores únicos y casa con el campo `fuente` del Contrato 1 (ruta relativa
POSIX). Verificado contra el corpus en disco: las 1.826 claves casan de forma
exacta con un archivo real, y 0 quedan sin archivo.

IDENTIDAD, NO SOLO COBERTURA
Desde el Q&A con los organizadores, la columna DOC_ID de este Excel ES el
`doc_id` del pipeline: el emparejamiento de la evaluación se hace con ese
valor, no con `fuente` ni con un hash del equipo. Por eso este módulo dejó de
ser un verificador de cobertura y pasó a ser la fuente de identidad.

La consecuencia práctica: aquí NO se puede devolver un diccionario vacío ni
tragarse un error. Un fallo silencioso deja todo el corpus sin `doc_id` y la
verificación de completitud da por buena una ingesta vacía. Todas las
validaciones de abajo lanzan excepción a propósito.
"""

from __future__ import annotations

from pathlib import Path

COL_CARPETA = "Carpeta"
COL_NOMBRE = "Nombre estandarizado"
COL_DOC_ID = "DOC_ID"
COL_FENOMENO = "Fenómeno"
COL_OBSERVATORIO = "Observatorio"
COL_CODIGO_OBS = "Código Observatorio"
HOJA = "Inventario de Archivos"


def _normalizar_ruta(carpeta, nombre) -> str:
    """Compone la ruta relativa POSIX tal como aparece en `fuente`."""
    carpeta = str(carpeta or "").strip().replace("\\", "/").strip("/")
    nombre = str(nombre or "").strip()
    return f"{carpeta}/{nombre}" if carpeta else nombre


def cargar_indice_oficial(ruta_xlsx: Path) -> dict[str, dict]:
    """
    Devuelve {ruta_relativa_posix: {doc_id_oficial, observatorio, ...}}.

    La clave es la ruta completa, NO el nombre de archivo: ver el docstring
    del módulo. Lanza ValueError si el resultado tiene menos entradas que
    filas el Excel, porque eso significa que la clave vuelve a colisionar.

    Raises:
        FileNotFoundError: no existe el XLSX.
        ValueError: falta la hoja, falta una columna, hay un DOC_ID vacío,
            hay DOC_ID repetidos, o la clave de ruta colisiona. Todos estos
            casos son fatales: sin índice íntegro no hay identidad.
    """
    import openpyxl

    ruta_xlsx = Path(ruta_xlsx)
    if not ruta_xlsx.exists():
        raise FileNotFoundError(
            f"No se encuentra el índice oficial en {ruta_xlsx}. Es la fuente "
            f"del doc_id: sin él la ingesta no puede identificar documentos."
        )

    wb = openpyxl.load_workbook(ruta_xlsx, read_only=True, data_only=True)
    try:
        # wb[HOJA] lanza un KeyError críptico si el nombre no está.
        if HOJA not in wb.sheetnames:
            raise ValueError(
                f"El índice {ruta_xlsx.name} no tiene la hoja {HOJA!r}. "
                f"Hojas encontradas: {wb.sheetnames}"
            )
        ws = wb[HOJA]
        filas = ws.iter_rows(values_only=True)
        cabecera = [str(c).strip() if c is not None else "" for c in next(filas)]
        idx = {nombre: i for i, nombre in enumerate(cabecera)}

        faltan = [c for c in (COL_CARPETA, COL_NOMBRE, COL_DOC_ID) if c not in idx]
        if faltan:
            raise ValueError(
                f"El índice no tiene las columnas esperadas: {faltan}. "
                f"Columnas encontradas: {cabecera}"
            )

        indice: dict[str, dict] = {}
        n_filas = 0
        for fila in filas:
            if fila is None or all(c is None for c in fila):
                continue
            n_filas += 1
            ruta = _normalizar_ruta(fila[idx[COL_CARPETA]], fila[idx[COL_NOMBRE]])
            if not ruta:
                continue

            # El DOC_ID es la identidad del documento y viaja hasta
            # resultados.jsonl. Un vacío aquí no se detecta más abajo: se
            # propagaría como null y el documento quedaría inevaluable.
            doc_id = str(fila[idx[COL_DOC_ID]] or "").strip()
            if not doc_id:
                raise ValueError(
                    f"DOC_ID vacío en la fila {n_filas + 1} del índice "
                    f"(ruta {ruta!r}). El DOC_ID es la identidad del documento: "
                    f"no puede faltar."
                )

            indice[ruta] = {
                "doc_id_oficial": doc_id,
                "nombre_archivo": str(fila[idx[COL_NOMBRE]] or "").strip(),
                "observatorio": fila[idx[COL_OBSERVATORIO]] if COL_OBSERVATORIO in idx else None,
                "codigo_observatorio": fila[idx[COL_CODIGO_OBS]] if COL_CODIGO_OBS in idx else None,
                "fenomeno_indice": fila[idx[COL_FENOMENO]] if COL_FENOMENO in idx else None,
            }
    finally:
        wb.close()

    if len(indice) < n_filas:
        raise ValueError(
            f"La clave del índice colisiona: {n_filas} filas -> {len(indice)} "
            f"entradas. Revisar _normalizar_ruta()."
        )

    # Simetría con la validación de arriba: hasta ahora se comprobaba que dos
    # filas no compartieran RUTA, pero no que no compartieran DOC_ID. Dos rutas
    # con el mismo DOC_ID producirían dos documentos con la misma identidad y
    # se pisarían en la evaluación.
    doc_ids = [v["doc_id_oficial"] for v in indice.values()]
    if len(set(doc_ids)) != len(doc_ids):
        repetidos = sorted({d for d in doc_ids if doc_ids.count(d) > 1})
        raise ValueError(
            f"DOC_ID repetidos en el índice: {repetidos[:10]}"
            f"{' ...' if len(repetidos) > 10 else ''} "
            f"({len(doc_ids)} filas, {len(set(doc_ids))} DOC_ID distintos)."
        )

    return indice


def comparar_con_ingesta(indice: dict[str, dict], fuentes_vistas: set[str]) -> dict:
    """
    Cruza el índice oficial con las `fuente` que la ingesta ha visto.

    Returns:
        faltan_por_ingerir: en el índice pero no vistas -> huecos de cobertura
        no_en_indice      : vistas pero no en el índice -> archivos extra en
                            el corpus, o rutas mal compuestas
    """
    del_indice = set(indice)
    return {
        "total_indice": len(del_indice),
        "total_vistas": len(fuentes_vistas),
        "cubiertas": len(del_indice & fuentes_vistas),
        "faltan_por_ingerir": sorted(del_indice - fuentes_vistas),
        "no_en_indice": sorted(fuentes_vistas - del_indice),
    }


if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) < 2:
        print("Uso: python indice_oficial.py <Indice_Datos_Codefest.xlsx> "
              "[documents.jsonl]")
        sys.exit(1)

    indice = cargar_indice_oficial(Path(sys.argv[1]))
    print(f"Índice oficial cargado: {len(indice)} archivos registrados")

    nombres = {v["nombre_archivo"] for v in indice.values()}
    print(f"  (nombres de archivo distintos: {len(nombres)} — "
          f"{len(indice) - len(nombres)} colisionarían si se indexara por nombre)")

    if len(sys.argv) > 2:
        fuentes = set()
        with open(sys.argv[2], encoding="utf-8") as f:
            for linea in f:
                if linea.strip():
                    fuentes.add(json.loads(linea)["fuente"])
        r = comparar_con_ingesta(indice, fuentes)
        print(f"\n  en el índice        : {r['total_indice']}")
        print(f"  vistas por la ingesta: {r['total_vistas']}")
        print(f"  cubiertas            : {r['cubiertas']}")
        print(f"  faltan por ingerir   : {len(r['faltan_por_ingerir'])}")
        print(f"  no están en el índice: {len(r['no_en_indice'])}")
        for f_ in r["no_en_indice"][:10]:
            print(f"    + {f_}")